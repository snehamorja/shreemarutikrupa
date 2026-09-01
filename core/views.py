import os
import shutil
import datetime
from decimal import Decimal
from pathlib import Path

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.core.mail import get_connection, EmailMessage
from django.http import HttpResponse, Http404
from django.conf import settings
from django.db import transaction

# Model and Form imports
from core.models import (
    BrandingSettings, Office, OfficeCredentialHistory, UserProfile, PricingConfig, 
    ScaifeEntry, Inquiry, FAQ, Testimonial, GalleryItem, ActivityLog, log_action,
    WorkerProfile, WorkerFinancialEntry
)
from core.forms import (
    ScaifeEntryForm, OfficeForm, PricingConfigForm, 
    BrandingSettingsForm, InquiryForm, UserCreationForm, UserEditForm,
    WorkerProfileForm, WorkerFinancialEntryForm
)

# OpenPyXL and ReportLab imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# ==========================================================================
# CUSTOM PERMISSION DECORATORS & HELPERS
# ==========================================================================
def admin_only(view_func):
    """
    Decorator to restrict view access to Super Admins only.
    """
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.profile.role != 'admin':
            messages.error(request, "Access denied. Administrator privileges required.")
            return redirect('overview')
        return view_func(request, *args, **kwargs)
    return _wrapped_view_func


def manager_or_admin(view_func):
    """
    Decorator to restrict view access to Managers and Super Admins.
    """
    def _wrapped_view_func(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.profile.role not in ['admin', 'manager']:
            messages.error(request, "Access denied. Manager privileges required.")
            return redirect('overview')
        return view_func(request, *args, **kwargs)
    return _wrapped_view_func


def send_inquiry_email(inquiry, branding):
    """
    Sends notification email for a new contact form submission.
    """
    subject = f"New Website Inquiry: {inquiry.subject}"
    body = f"""
New Inquiry Received from Public Website:

Name: {inquiry.name}
Email: {inquiry.email}
Phone: {inquiry.phone}
Subject: {inquiry.subject}

Message:
{inquiry.message}

--
Manage this inquiry in your Portal Dashboard.
"""
    try:
        if branding.smtp_host and branding.smtp_user:
            connection = get_connection(
                backend='django.core.mail.backends.smtp.EmailBackend',
                host=branding.smtp_host,
                port=branding.smtp_port,
                username=branding.smtp_user,
                password=branding.smtp_password,
                use_tls=branding.smtp_use_tls,
                use_ssl=branding.smtp_use_ssl,
            )
            email = EmailMessage(
                subject=subject,
                body=body,
                from_email=branding.smtp_from_email or branding.contact_email,
                to=[branding.contact_email],
                connection=connection
            )
            email.send()
        else:
            # Fallback to default setting (console backend during dev)
            from django.core.mail import send_mail
            send_mail(
                subject,
                body,
                branding.contact_email,
                [branding.contact_email],
                fail_silently=True
            )
    except Exception as e:
        print(f"Failed to send email notification: {e}")


# ==========================================================================
# PUBLIC PAGES VIEWS
# ==========================================================================
def home(request):
    if request.user.is_authenticated:
        return redirect('overview')
    return redirect('login')


def about(request):
    return render(request, 'about.html')


def services(request):
    return render(request, 'services.html')


def gallery(request):
    items = GalleryItem.objects.all()
    return render(request, 'gallery.html', {'items': items})


def faqs(request):
    faqs_list = FAQ.objects.filter(is_active=True)
    return render(request, 'faq.html', {'faqs': faqs_list})


def terms(request):
    return render(request, 'terms.html')


def privacy(request):
    return render(request, 'privacy.html')


def contact(request):
    if request.method == 'POST':
        form = InquiryForm(request.POST)
        if form.is_valid():
            inquiry = form.save()
            branding = BrandingSettings.get_solo()
            
            # Send Notification Email
            send_inquiry_email(inquiry, branding)
            
            log_action(
                user=None,
                action_type="CREATE",
                model_name="Inquiry",
                object_repr=f"Inquiry from {inquiry.name}",
                details=f"Inquiry regarding '{inquiry.subject}' registered successfully."
            )
            
            messages.success(request, "Your inquiry has been submitted successfully! We will contact you shortly.")
            return redirect('contact')
        else:
            messages.error(request, "Please fill out the contact form correctly.")
    return render(request, 'contact.html')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('overview')

    next_url = request.GET.get('next', 'overview')
    error_msg = None

    if request.method == 'POST':
        u_input = (request.POST.get('username') or '').strip()
        p_input = (request.POST.get('password') or '').strip()
        next_url = request.POST.get('next', 'overview')

        # Allow login by username OR email address
        actual_username = u_input
        if '@' in u_input:
            user_by_email = User.objects.filter(email__iexact=u_input).first()
            if user_by_email:
                actual_username = user_by_email.username
        else:
            user_by_uname = User.objects.filter(username__iexact=u_input).first()
            if user_by_uname:
                actual_username = user_by_uname.username

        user = authenticate(request, username=actual_username, password=p_input)
        if user is not None:
            login(request, user)
            log_action(user, "LOGIN", "User", user.username, "User logged in to dashboard.")
            messages.success(request, f"Welcome back, {user.first_name or user.username}!")
            return redirect(next_url)
        else:
            error_msg = "Invalid username/email or password. Please check your credentials."

    return render(request, 'login.html', {'error_msg': error_msg, 'next': next_url})


def logout_view(request):
    if request.user.is_authenticated:
        log_action(request.user, "LOGOUT", "User", request.user.username, "User logged out.")
        logout(request)
        messages.success(request, "You have been logged out successfully.")
    return redirect('login')


# ==========================================================================
# DASHBOARD CORE VIEWS
# ==========================================================================
@login_required
def overview(request):
    user_profile = request.user.profile
    is_admin = user_profile.role == 'admin'

    # Determine scaife queryset based on role & office
    if is_admin:
        scaifes = ScaifeEntry.objects.all()
        office_label = "All Offices (Global View)"
        office_obj = None
    elif user_profile.office:
        scaifes = ScaifeEntry.objects.filter(assigned_office=user_profile.office)
        office_label = user_profile.office.name
        office_obj = user_profile.office
    else:
        scaifes = ScaifeEntry.objects.none()
        office_label = "Unassigned"
        office_obj = None

    scaifes_total = sum(entry.quantity for entry in scaifes.exclude(status='delivered'))  # sum of scaife quantities
    scaifes_processing = sum(entry.quantity for entry in scaifes.exclude(status__in=['delivered', 'ready']))
    scaifes_delivered = sum(entry.quantity for entry in scaifes.filter(status='delivered'))
    scaifes_ready = sum(entry.quantity for entry in scaifes.filter(status='ready'))

    # Revenue: total cost of ALL scaifes for this office
    total_revenue = sum(entry.cost for entry in scaifes)

    # Inquiries: admin sees all, others see none (contact form is public)
    if is_admin:
        inquiries_new = Inquiry.objects.filter(status='new').count()
        recent_inquiries = Inquiry.objects.all().order_by('-created_at')[:5]
    else:
        inquiries_new = None  # not shown to non-admins
        recent_inquiries = []

    # Recent active scaifes (office-filtered)
    recent_scaifes = scaifes.exclude(status='delivered').order_by('-received_date')[:8]

    # Chart 1: Status distribution for this office
    status_map = {'received': 0, 'lapping': 0, 'coating': 0, 'qc': 0, 'ready': 0, 'delivered': 0}
    for item in scaifes:
        if item.status in status_map:
            status_map[item.status] += item.quantity

    status_labels = ["Received", "In Lapping", "In Coating", "QC", "Ready", "Delivered"]
    status_counts = [
        status_map['received'], status_map['lapping'], status_map['coating'],
        status_map['qc'], status_map['ready'], status_map['delivered']
    ]

    # Chart 2: Monthly scaife count for this office (last 6 months)
    months_labels = []
    monthly_counts = []
    monthly_revenue = []
    today = datetime.date.today()
    for i in range(5, -1, -1):
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        months_labels.append(datetime.date(year, month, 1).strftime("%b %Y"))
        month_qs = scaifes.filter(received_date__year=year, received_date__month=month)
        monthly_counts.append(sum(e.quantity for e in month_qs))
        monthly_revenue.append(float(sum(e.cost for e in month_qs)))

    # Per-office breakdown (admin only)
    office_breakdown = []
    if is_admin:
        for office in Office.objects.all():
            o_qs = ScaifeEntry.objects.filter(assigned_office=office)
            office_breakdown.append({
                'id': office.id,
                'name': office.name,
                'location': office.location,
                'total': sum(e.quantity for e in o_qs),
                'revenue': sum(e.cost for e in o_qs),
            })

    context = {
        'scaifes_total': scaifes_total,
        'scaifes_processing': scaifes_processing,
        'scaifes_delivered': scaifes_delivered,
        'scaifes_ready': scaifes_ready,
        'total_revenue': total_revenue,
        'inquiries_new': inquiries_new,
        'recent_scaifes': recent_scaifes,
        'recent_inquiries': recent_inquiries,
        'status_labels': status_labels,
        'status_counts': status_counts,
        'months_labels': months_labels,
        'monthly_counts': monthly_counts,
        'monthly_revenue': monthly_revenue,
        'office_label': office_label,
        'office_obj': office_obj,
        'is_admin': is_admin,
        'office_breakdown': office_breakdown,
    }
    return render(request, 'dashboard/overview.html', context)


@login_required
def scaife_list(request):
    user_profile = request.user.profile
    
    # Active scaifes show items received within the last 24 hours (today & yesterday)
    import datetime
    from django.utils import timezone
    cutoff_24h = timezone.now().date() - datetime.timedelta(days=1)

    if user_profile.role == 'admin':
        base_queryset = ScaifeEntry.objects.filter(received_date__gte=cutoff_24h).order_by('-received_date', '-id')
    else:
        base_queryset = ScaifeEntry.objects.filter(assigned_office=user_profile.office, received_date__gte=cutoff_24h).order_by('-received_date', '-id')

    # Apply search and dropdown filters
    query = request.GET.get('q', '')
    filter_status = request.GET.get('status', '')
    filter_office = request.GET.get('office', '')
    filter_service = request.GET.get('service', '')

    queryset = base_queryset
    if query:
        queryset = queryset.filter(client_name__icontains=query) | queryset.filter(serial_number__icontains=query)
    if filter_status:
        queryset = queryset.filter(status=filter_status)
    if filter_office:
        queryset = queryset.filter(assigned_office_id=filter_office)

    # Apply service filter (includes entries that have this service, even if they have multiple)
    if filter_service == 'lapping':
        queryset = queryset.filter(service_lapping=True)
    elif filter_service == 'coating':
        queryset = queryset.filter(service_coating=True)
    elif filter_service == 'diamond':
        queryset = queryset.filter(service_diamond_scaife=True)

    # Summary KPI totals (sum of quantities for all active/non-delivered scaifes)
    active_qs = base_queryset.exclude(status='delivered')
    total_active_scaifes = sum(e.quantity for e in active_qs)
    lapping_total = sum(e.quantity for e in active_qs if e.service_lapping)
    coating_total = sum(e.quantity for e in active_qs if e.service_coating)
    diamond_total = sum(e.quantity for e in active_qs if e.service_diamond_scaife)

    # Pagination
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    offices = Office.objects.all()

    is_admin = user_profile.role == 'admin'
    office_obj = user_profile.office if not is_admin else None
    office_label = office_obj.name if office_obj else 'All Offices'

    # Check for auto_wp / notify_wp param — build WhatsApp notification
    notify_entry = None
    notify_wp_url = None
    auto_open_wp = False

    branding = BrandingSettings.get_solo()
    admin_wp = (branding.whatsapp_number or '').strip().replace('+', '').replace(' ', '')

    # auto_wp = auto-open WhatsApp immediately on page load (after create/edit)
    # notify_wp = show banner only (manual click)
    wp_entry_id = request.GET.get('auto_wp') or request.GET.get('notify_wp')
    auto_open_wp = bool(request.GET.get('auto_wp'))

    if wp_entry_id:
        notify_entry = ScaifeEntry.objects.filter(pk=wp_entry_id).first()
        if notify_entry:
            from urllib.parse import quote
            office_name = notify_entry.assigned_office.name if notify_entry.assigned_office else 'Factory'
            repair_text = f'⚠️ REPAIR NEEDED: {notify_entry.repair_details}' if notify_entry.needs_repair else ''
            services = notify_entry.get_services_display()
            msg_parts = [
                f"🏭 *{office_name}*",
                f"📦 Scaife Entry — {notify_entry.received_date.strftime('%d %b %Y')}",
                f"🔧 Services: {services}",
                f"📊 Quantity: {notify_entry.quantity}",
            ]
            if notify_entry.service_lapping and notify_entry.quantity_lapping:
                msg_parts.append(f"  • Lapping: {notify_entry.quantity_lapping} pcs")
            if notify_entry.service_coating and notify_entry.quantity_coating:
                msg_parts.append(f"  • Coating: {notify_entry.quantity_coating} pcs")
            if notify_entry.service_diamond_scaife and notify_entry.quantity_diamond:
                msg_parts.append(f"  • Diamond Scaife: {notify_entry.quantity_diamond} pcs")
            if repair_text:
                msg_parts.append(repair_text)
            msg_parts.append(f"💰 Total Bill: ₹{notify_entry.cost:.2f}")
            msg = "\n".join(msg_parts)
            encoded_msg = quote(msg)
            notify_wp_url = f"https://wa.me/{admin_wp}?text={encoded_msg}" if admin_wp else f"https://wa.me/?text={encoded_msg}"

    context = {
        'scaife_entries': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'offices': offices,
        'query': query,
        'filter_status': filter_status,
        'filter_office': filter_office,
        'filter_service': filter_service,
        'is_admin': is_admin,
        'office_obj': office_obj,
        'office_label': office_label,
        'total_active_scaifes': total_active_scaifes,
        'lapping_total': lapping_total,
        'coating_total': coating_total,
        'diamond_total': diamond_total,
        'notify_entry': notify_entry,
        'notify_wp_url': notify_wp_url,
        'auto_open_wp': auto_open_wp,
        'admin_wp': admin_wp,
    }
    return render(request, 'dashboard/scaife_list.html', context)


@login_required
def scaife_history(request):
    """Historical scaife data — filter by month, year, or date range."""
    user_profile = request.user.profile

    if user_profile.role == 'admin':
        queryset = ScaifeEntry.objects.all().order_by('-received_date')
    else:
        queryset = ScaifeEntry.objects.filter(assigned_office=user_profile.office).order_by('-received_date')

    # Get filter params
    query = request.GET.get('q', '')
    filter_status = request.GET.get('status', '')
    filter_office = request.GET.get('office', '')
    filter_year = request.GET.get('year', '')
    filter_month = request.GET.get('month', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')
    filter_service = request.GET.get('service', '')  # lapping | coating | diamond

    # Apply year filter
    if filter_year:
        queryset = queryset.filter(received_date__year=int(filter_year))

    # Apply month filter (works with or without year)
    if filter_month:
        queryset = queryset.filter(received_date__month=int(filter_month))

    # Apply date range filter (overrides year/month if both provided)
    if date_from:
        queryset = queryset.filter(received_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(received_date__lte=date_to)

    # Apply text search
    if query:
        queryset = queryset.filter(client_name__icontains=query) | queryset.filter(serial_number__icontains=query)
    if filter_status:
        queryset = queryset.filter(status=filter_status)
    if filter_office:
        queryset = queryset.filter(assigned_office_id=filter_office)
    # Apply service filter
    if filter_service == 'lapping':
        queryset = queryset.filter(service_lapping=True)
    elif filter_service == 'coating':
        queryset = queryset.filter(service_coating=True)
    elif filter_service == 'diamond':
        queryset = queryset.filter(service_diamond_scaife=True)

    # Summary stats for filtered results
    total_count = queryset.count()
    total_revenue = sum(e.cost for e in queryset)
    delivered_count = queryset.filter(status='delivered').count()

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    offices = Office.objects.all()
    is_admin = user_profile.role == 'admin'
    office_obj = user_profile.office if not is_admin else None
    office_label = office_obj.name if office_obj else 'All Offices'

    # Build available years from data
    all_years = ScaifeEntry.objects.dates('received_date', 'year', order='DESC')
    available_years = [d.year for d in all_years]

    context = {
        'scaife_entries': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'offices': offices,
        'query': query,
        'filter_status': filter_status,
        'filter_office': filter_office,
        'filter_year': filter_year,
        'filter_month': filter_month,
        'date_from': date_from,
        'date_to': date_to,
        'filter_service': filter_service,
        'is_admin': is_admin,
        'office_obj': office_obj,
        'office_label': office_label,
        'available_years': available_years,
        'total_count': total_count,
        'total_revenue': total_revenue,
        'delivered_count': delivered_count,
    }
    return render(request, 'dashboard/scaife_history.html', context)


@login_required
@manager_or_admin
def scaife_create(request):
    user_profile = request.user.profile
    pricing_config = PricingConfig.get_solo()

    if request.method == 'POST':
        form = ScaifeEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.created_by = request.user

            # Auto-assign office based on current user profile
            if user_profile.office:
                entry.assigned_office = user_profile.office
            else:
                entry.assigned_office = Office.objects.first()

            # Auto-set client_name from office name (no manual entry needed)
            entry.client_name = entry.assigned_office.name if entry.assigned_office else 'Factory'
            entry.status = 'received'
            entry.save()

            log_action(
                user=request.user,
                action_type="CREATE",
                model_name="ScaifeEntry",
                object_repr=entry.client_name,
                details=f"Scaife entry for '{entry.client_name}' registered (Qty: {entry.quantity}). Services: {entry.get_services_display()}. Bill: ₹{entry.cost}"
            )
            messages.success(request, f"✅ Scaife entry saved for '{entry.client_name}' — Qty: {entry.quantity}, Bill: ₹{entry.cost:.2f}")
            return redirect(f"/dashboard/scaife/?auto_wp={entry.pk}")
        else:
            messages.error(request, "Failed to create entry. Please inspect the form input values.")
    else:
        form = ScaifeEntryForm(initial={'quantity': 1, 'service_lapping': True})

    context = {
        'form': form,
        'pricing_config': pricing_config,
        'office_name': user_profile.office.name if user_profile.office else 'Unassigned',
    }
    return render(request, 'dashboard/scaife_form.html', context)


@login_required
def scaife_edit(request, pk):
    entry = get_object_or_404(ScaifeEntry, pk=pk)
    user_profile = request.user.profile
    pricing_config = PricingConfig.get_solo()

    if user_profile.role != 'admin' and user_profile.office and entry.assigned_office != user_profile.office:
        messages.error(request, "You do not have permission to edit scaife entries from other offices.")
        return redirect('scaife_list')

    if request.method == 'POST':
        form = ScaifeEntryForm(request.POST, instance=entry)
        if form.is_valid():
            entry = form.save(commit=False)
            # Keep client_name in sync with office
            if entry.assigned_office:
                entry.client_name = entry.assigned_office.name
            entry.save()
            log_action(
                user=request.user,
                action_type="UPDATE",
                model_name="ScaifeEntry",
                object_repr=entry.client_name,
                details=f"Scaife entry '{entry.client_name}' updated. Qty: {entry.quantity}, Bill: ₹{entry.cost}"
            )
            messages.success(request, f"✅ Scaife entry updated for '{entry.client_name}' — Bill: ₹{entry.cost:.2f}")
            return redirect(f"/dashboard/scaife/?auto_wp={entry.pk}")
    else:
        form = ScaifeEntryForm(instance=entry)

    context = {
        'form': form,
        'entry': entry,
        'pricing_config': pricing_config,
        'office_name': entry.assigned_office.name if entry.assigned_office else (user_profile.office.name if user_profile.office else 'Unassigned'),
    }
    return render(request, 'dashboard/scaife_form.html', context)


@login_required
@admin_only
def scaife_delete(request, pk):
    entry = get_object_or_404(ScaifeEntry, pk=pk)
    serial = entry.serial_number
    entry.delete()
    log_action(
        user=request.user,
        action_type="DELETE",
        model_name="ScaifeEntry",
        object_repr=serial,
        details=f"Scaife entry deleted by Admin."
    )
    messages.success(request, f"Scaife registry {serial} removed.")
    return redirect('scaife_list')


# ==========================================================================
# OFFICE MANAGEMENT CRUD
# ==========================================================================
@login_required
@manager_or_admin
def office_list(request):
    offices = Office.objects.all()

    # Annotate each office with its manager username (first manager found)
    for office in offices:
        manager_profile = UserProfile.objects.filter(office=office, role='manager').select_related('user').first()
        office.manager_username = manager_profile.user.username if manager_profile else None

    # Pop one-time credentials from session (shown after creating a new office)
    new_office_creds = request.session.pop('new_office_credentials', None)

    return render(request, 'dashboard/office_list.html', {
        'offices': offices,
        'new_office_creds': new_office_creds,
    })


@login_required
@admin_only
def office_create(request):
    if request.method == 'POST':
        form = OfficeForm(request.POST)
        
        custom_username = (request.POST.get('manager_username') or '').strip()
        custom_password = (request.POST.get('manager_password') or '').strip()

        # Validate username uniqueness if admin provided a custom username
        if custom_username and User.objects.filter(username__iexact=custom_username).exists():
            form.add_error('manager_username', f"Username '{custom_username}' is already taken. Please choose a different username.")

        if form.is_valid():
            office = form.save()

            # ── Determine Username ──
            if custom_username:
                username = custom_username
            else:
                # Auto-generate username from office name
                base_username = office.name.lower().replace(' ', '_')
                base_username = ''.join(c for c in base_username if c.isalnum() or c == '_')
                username = base_username
                counter = 1
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1

            # ── Determine Password ──
            if custom_password:
                password = custom_password
            else:
                import string, secrets
                alphabet = string.ascii_letters + string.digits
                password = ''.join(secrets.choice(alphabet) for _ in range(12))

            # Create Manager User Account
            new_user = User.objects.create_user(
                username=username,
                password=password,
                email=office.email,
                first_name=office.name,
                last_name='Manager',
            )
            new_user.profile.role = 'manager'
            new_user.profile.office = office
            new_user.profile.save()

            # Save username into Office model
            office.manager_username = username
            office.save()

            # Record credential history
            OfficeCredentialHistory.objects.create(
                office=office,
                username=username,
                password_text=password,
                changed_by=request.user,
                change_type="Manager Account Created"
            )

            # Send Email
            try:
                branding = BrandingSettings.objects.filter(pk=1).first()
                company = branding.company_name if branding else "Scaife Portal"

                subject = f"Your Dashboard Login Credentials — {company}"
                body = (
                    f"Hello {office.name} Team,\n\n"
                    f"Your office has been registered on the {company} portal.\n"
                    f"Below are your manager dashboard login credentials:\n\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"  🔑 Login URL  : {request.build_absolute_uri('/login/')}\n"
                    f"  👤 Username   : {username}\n"
                    f"  🔒 Password   : {password}\n"
                    f"  🏢 Office     : {office.name}\n"
                    f"  📍 Location   : {office.location}\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    f"Best regards,\n{company} Admin Team"
                )

                email_msg = EmailMessage(
                    subject=subject,
                    body=body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[office.email],
                )
                email_msg.send(fail_silently=True)
            except Exception:
                pass

            log_action(
                user=request.user,
                action_type="CREATE",
                model_name="Office",
                object_repr=office.name,
                details=f"Office '{office.name}' created with manager username '{username}'."
            )

            messages.success(
                request,
                f"✅ Office '{office.name}' created successfully! "
                f"Manager Username: {username} | Password: {password}"
            )

            request.session['new_office_credentials'] = {
                'office_name': office.name,
                'username': username,
                'password': password,
                'email': office.email,
                'phone': office.phone,
            }

            return redirect('office_list')
    else:
        form = OfficeForm()
    return render(request, 'dashboard/office_form.html', {'form': form})


@login_required
@admin_only
def office_edit(request, pk):
    office = get_object_or_404(Office, pk=pk)

    # Fetch the manager user linked to this office
    manager_profile = UserProfile.objects.filter(office=office, role='manager').select_related('user').first()
    manager_user = manager_profile.user if manager_profile else None

    reset_creds = None

    if request.method == 'POST':
        action = request.POST.get('action', 'save')

        if action == 'reset_password' and manager_user:
            # Check for custom password choice by admin
            custom_pwd = (request.POST.get('custom_new_password') or '').strip()
            if custom_pwd:
                new_password = custom_pwd
            else:
                import string, secrets
                alphabet = string.ascii_letters + string.digits
                new_password = ''.join(secrets.choice(alphabet) for _ in range(12))

            manager_user.set_password(new_password)
            manager_user.save()

            # Email new password
            try:
                branding = BrandingSettings.objects.filter(pk=1).first()
                company = branding.company_name if branding else "Scaife Portal"
                from django.core.mail import EmailMessage as EmailMsg
                EmailMsg(
                    subject=f"Password Reset — {company}",
                    body=(
                        f"Hello {office.name} Team,\n\n"
                        f"Your dashboard password has been reset by the admin.\n\n"
                        f"  👤 Username : {manager_user.username}\n"
                        f"  🔒 New Password : {new_password}\n"
                        f"  🔑 Login URL : {request.build_absolute_uri('/login/')}\n\n"
                        f"Please change your password after logging in.\n\nRegards,\n{company} Admin"
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[office.email],
                ).send(fail_silently=True)
            except Exception:
                pass

            # Record credential history
            OfficeCredentialHistory.objects.create(
                office=office,
                username=manager_user.username,
                password_text=new_password,
                changed_by=request.user,
                change_type="Password Changed by Admin" if custom_pwd else "Password Reset by Admin"
            )

            log_action(request.user, "UPDATE", "Office", office.name,
                       f"Manager password updated for '{manager_user.username}'.")
            messages.success(request, f"Password updated for '{manager_user.username}'. New credentials shown below.")
            reset_creds = {
                'username': manager_user.username,
                'password': new_password,
                'email': office.email,
            }
            # Re-render same page with new creds
            form = OfficeForm(instance=office)
            latest_password = reset_creds['password']
            return render(request, 'dashboard/office_form.html', {
                'form': form, 'office': office,
                'manager_user': manager_user, 'reset_creds': reset_creds,
                'latest_password': latest_password,
                'credential_history': office.credential_history.all(),
            })

        else:
            form = OfficeForm(request.POST, instance=office)
            if form.is_valid():
                form.save()
                log_action(request.user, "UPDATE", "Office", office.name,
                           "Office branch parameters updated.")
                messages.success(request, f"Office {office.name} updated successfully.")
                return redirect('office_list')
    else:
        form = OfficeForm(instance=office)

    latest_creds = office.credential_history.first()
    latest_password = latest_creds.password_text if latest_creds else ''

    return render(request, 'dashboard/office_form.html', {
        'form': form, 'office': office,
        'manager_user': manager_user,
        'latest_password': latest_password,
        'credential_history': office.credential_history.all(),
    })


@login_required
@admin_only
def office_delete(request, pk):
    office = get_object_or_404(Office, pk=pk)
    name = office.name
    office.delete()
    log_action(
        user=request.user,
        action_type="DELETE",
        model_name="Office",
        object_repr=name,
        details=f"Office center removed by Admin."
    )
    messages.success(request, f"Office {name} deleted successfully.")
    return redirect('office_list')


# ==========================================================================
# PRICING CONFIGURATION View/Edit
# ==========================================================================
@login_required
@admin_only
def pricing_edit(request):
    pricing = PricingConfig.get_solo()
    if request.method == 'POST':
        form = PricingConfigForm(request.POST, instance=pricing)
        if form.is_valid():
            form.save()
            log_action(
                user=request.user,
                action_type="UPDATE",
                model_name="PricingConfig",
                object_repr="Pricing Configuration",
                details="Lapping/Coating rate constants altered."
            )
            messages.success(request, "Pricing configuration updated successfully.")
            return redirect('overview')
    else:
        form = PricingConfigForm(instance=pricing)
    return render(request, 'dashboard/pricing_form.html', {'form': form})


# ==========================================================================
# USER PROFILE UPDATE (FIRST NAME, LAST NAME, EMAIL, PASSWORD)
# ==========================================================================
@login_required
def profile_edit(request):
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        # 1. Update personal details
        if form_type == 'profile_info':
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            username = request.POST.get('username')
            
            user = request.user
            if username and username != user.username:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' is already taken.")
                    return redirect('profile_edit')
                user.username = username
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.save()
            
            log_action(user, "UPDATE", "User", user.username, "Profile contact metadata details updated.")
            messages.success(request, "Profile information updated successfully.")
            
        # 2. Reset password security
        elif form_type == 'password_change':
            current_pass = request.POST.get('current_password')
            new_pass = request.POST.get('new_password')
            confirm_pass = request.POST.get('confirm_password')
            
            user = request.user
            if not user.check_password(current_pass):
                messages.error(request, "Failed: Your current password was entered incorrectly.")
            elif new_pass != confirm_pass:
                messages.error(request, "Failed: New password configurations do not match.")
            else:
                user.set_password(new_pass)
                user.save()
                # Re-authenticate because password change expires the session
                login(request, user)
                log_action(user, "UPDATE", "User", user.username, "Login password reset by user.")

                # Record credential history if user is linked to an office
                if hasattr(user, 'profile') and user.profile.office:
                    OfficeCredentialHistory.objects.create(
                        office=user.profile.office,
                        username=user.username,
                        password_text=new_pass,
                        changed_by=user,
                        change_type="Password Changed by User (Profile)"
                    )

                messages.success(request, "Your password has been changed successfully.")
        
        return redirect('profile_edit')
        
    return render(request, 'dashboard/profile.html')


# ==========================================================================
# SYSTEM USER ACCOUNTS CRUD (Super Admin Only)
# ==========================================================================
@login_required
@admin_only
def user_list(request):
    profiles = UserProfile.objects.all().select_related('user', 'office')
    return render(request, 'dashboard/user_list.html', {'profiles': profiles})


@login_required
@admin_only
def user_create(request):
    offices = Office.objects.all()
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            new_user = form.save()
            log_action(
                user=request.user,
                action_type="CREATE",
                model_name="UserProfile",
                object_repr=new_user.username,
                details=f"New portal account '{new_user.username}' created. Role: {new_user.profile.role}."
            )
            messages.success(request, f"User {new_user.username} registered successfully.")
            return redirect('user_list')
    else:
        form = UserCreationForm()
    return render(request, 'dashboard/user_form.html', {'form': form, 'offices': offices, 'edit_mode': False})


@login_required
@admin_only
def user_edit(request, pk):
    target_user = get_object_or_404(User, pk=pk)
    offices = Office.objects.all()
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=target_user)
        if form.is_valid():
            form.save()
            log_action(
                user=request.user,
                action_type="UPDATE",
                model_name="UserProfile",
                object_repr=target_user.username,
                details=f"Account configuration settings updated."
            )
            messages.success(request, f"User {target_user.username} updated successfully.")
            return redirect('user_list')
    else:
        form = UserEditForm(instance=target_user)
    
    context = {
        'form': form,
        'offices': offices,
        'edit_mode': True,
        'role_value': target_user.profile.role,
        'office_value': target_user.profile.office.id if target_user.profile.office else None
    }
    return render(request, 'dashboard/user_form.html', context)


@login_required
@admin_only
def user_delete(request, pk):
    target_user = get_object_or_404(User, pk=pk)
    username = target_user.username
    if target_user.username == 'admin' or target_user.pk == request.user.pk:
        messages.error(request, "Failed: You cannot delete your active admin session profile.")
    else:
        target_user.delete()
        log_action(
            user=request.user,
            action_type="DELETE",
            model_name="UserProfile",
            object_repr=username,
            details=f"User operator credentials removed."
        )
        messages.success(request, f"User {username} deleted successfully.")
    return redirect('user_list')


# ==========================================================================
# AUDIT LOGS DISPLAY
# ==========================================================================
@login_required
@admin_only
def activity_logs(request):
    logs = ActivityLog.objects.all().select_related('user', 'user__profile')[:100]
    return render(request, 'dashboard/activity_log.html', {'logs': logs})


# ==========================================================================
# BRANDING SETTINGS & DATABASE BACKUPS View
# ==========================================================================
@login_required
@admin_only
def settings_edit(request):
    branding = BrandingSettings.get_solo()
    
    # Handle Branding Settings Post
    if request.method == 'POST' and request.POST.get('form_type') == 'branding':
        form = BrandingSettingsForm(request.POST, request.FILES, instance=branding)
        if form.is_valid():
            # Server-side 2-step PIN confirmation check
            new_pin = request.POST.get('admin_security_pin', '').strip()
            confirm_pin = request.POST.get('admin_security_pin_confirm', '').strip()
            if confirm_pin and new_pin != confirm_pin:
                messages.error(request, "❌ Master Security PINs do not match. Please enter both PINs correctly and try again.")
                form_obj = BrandingSettingsForm(instance=branding)
                return render(request, 'dashboard/settings.html', {'form': form_obj, 'branding': branding, 'backup_files': []})
            # Handle manual toggle checks due to HTML value anomalies
            settings_obj = form.save(commit=False)
            settings_obj.smtp_use_tls = 'smtp_use_tls' in request.POST
            settings_obj.smtp_use_ssl = 'smtp_use_ssl' in request.POST
            settings_obj.save()
            
            log_action(
                user=request.user,
                action_type="UPDATE",
                model_name="BrandingSettings",
                object_repr="Branding Settings",
                details="Portal visual identity, mapping configs, or SMTP setups modified."
            )
            messages.success(request, "Branding and portal settings updated successfully.")
            return redirect('settings_edit')
        else:
            messages.error(request, "Failed to update settings. Verify parameters.")
    
    # Backups directory setups
    backup_dir = Path(settings.BASE_DIR) / 'backups'
    backup_dir.mkdir(exist_ok=True)
    
    # List available backup files
    backup_files = []
    for f in backup_dir.glob("*.sqlite3"):
        stat = f.stat()
        backup_files.append({
            'filename': f.name,
            'size': stat.st_size,
            'created': datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%d %b %Y %H:%M:%S")
        })
    # Order latest backups first
    backup_files.sort(key=lambda x: x['filename'], reverse=True)

    form = BrandingSettingsForm(instance=branding)
    context = {
        'form': form,
        'branding': branding,
        'backup_files': backup_files
    }
    return render(request, 'dashboard/settings.html', context)


# ==========================================================================
# DATABASE SQLITE BACKUP & RESTORE VIEW UTILITIES
# ==========================================================================
@login_required
@admin_only
def backup_create(request):
    try:
        backup_dir = Path(settings.BASE_DIR) / 'backups'
        backup_dir.mkdir(exist_ok=True)
        
        # Backup file formatting
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.sqlite3"
        backup_path = backup_dir / backup_filename
        
        # SQLite copy file
        src_path = Path(settings.BASE_DIR) / 'db.sqlite3'
        shutil.copy2(src_path, backup_path)
        
        log_action(
            user=request.user,
            action_type="CREATE",
            model_name="DatabaseBackup",
            object_repr=backup_filename,
            details=f"SQLite database snapshot file successfully created and downloaded."
        )

        # Auto-download the backup file immediately
        with open(backup_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/x-sqlite3")
            response['Content-Disposition'] = f'attachment; filename={backup_filename}'
            return response

    except Exception as e:
        messages.error(request, f"Error generating database backup: {e}")
        
    return redirect('settings_edit')


@login_required
@admin_only
def backup_download(request, filename):
    backup_dir = Path(settings.BASE_DIR) / 'backups'
    file_path = backup_dir / filename
    
    if file_path.exists() and file_path.name.endswith('.sqlite3'):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/x-sqlite3")
            response['Content-Disposition'] = f'attachment; filename={file_path.name}'
            return response
    raise Http404("Backup file not found.")


@login_required
@admin_only
def backup_restore(request):
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        
        # Validate that the file is an SQLite file by filename extension
        if not backup_file.name.endswith('.sqlite3'):
            messages.error(request, "Failed: Invalid file extension. Only .sqlite3 backup archives are allowed.")
            return redirect('settings_edit')
            
        active_db_path = Path(settings.BASE_DIR) / 'db.sqlite3'
        temp_rollback_path = Path(settings.BASE_DIR) / 'db_rollback_temp.sqlite3'
        
        # Step 1: Create a temp backup copy of the current active DB (safe-guard)
        try:
            if active_db_path.exists():
                shutil.copy2(active_db_path, temp_rollback_path)
        except Exception as e:
            messages.error(request, f"Rollback preparation error: {e}")
            return redirect('settings_edit')
            
        # Step 2: Overwrite the active DB with the uploaded file
        try:
            # Open target file block and overwrite
            with open(active_db_path, 'wb+') as destination:
                for chunk in backup_file.chunks():
                    destination.write(chunk)
                    
            # Step 3: Run quick test check query to ensure database integrity is intact
            # We fetch user list count to trigger sqlite test query
            list(User.objects.all())
            
            # Clean up temp file
            if temp_rollback_path.exists():
                os.remove(temp_rollback_path)
                
            log_action(
                user=request.user,
                action_type="RESTORE",
                model_name="Database",
                object_repr=backup_file.name,
                details=f"Database state successfully restored from backup file upload."
            )
            messages.success(request, f"Database restored successfully from '{backup_file.name}'!")
            
        except Exception as e:
            # Step 4: Rollback database to pre-upload state if exception occurs
            if temp_rollback_path.exists():
                shutil.copy2(temp_rollback_path, active_db_path)
                os.remove(temp_rollback_path)
            messages.error(request, f"Restoration failed! SQLite database corrupted or invalid: {e}. Rollback executed.")
            
    return redirect('settings_edit')


# ==========================================================================
# REPORT EXPORTING VIEWS (EXCEL & PDF GENERATION WITH FILTER RETENTION)
# ==========================================================================
def get_filtered_scaifes(request):
    """
    Helper function to filter queryset consistently for both Excel and PDF reporting.
    """
    user_profile = request.user.profile
    if user_profile.role == 'admin':
        queryset = ScaifeEntry.objects.all().order_by('-received_date')
    else:
        queryset = ScaifeEntry.objects.filter(assigned_office=user_profile.office).order_by('-received_date')

    query = request.GET.get('q', '')
    filter_status = request.GET.get('status', '')
    filter_office = request.GET.get('office', '')
    filter_year = request.GET.get('year', '')
    filter_month = request.GET.get('month', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    if filter_year:
        queryset = queryset.filter(received_date__year=int(filter_year))
    if filter_month:
        queryset = queryset.filter(received_date__month=int(filter_month))
    if date_from:
        queryset = queryset.filter(received_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(received_date__lte=date_to)

    if query:
        queryset = queryset.filter(client_name__icontains=query) | queryset.filter(serial_number__icontains=query)
    if filter_status:
        queryset = queryset.filter(status=filter_status)
    if filter_office:
        queryset = queryset.filter(assigned_office_id=filter_office)
        
    return queryset


@login_required
def export_excel(request):
    queryset = get_filtered_scaifes(request)
    
    # Create workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scaife Process Reports"
    
    # Font configs
    title_font = Font(name='Arial', size=16, bold=True, color='0EA5E9')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Arial', size=10, bold=True)
    regular_font = Font(name='Arial', size=10)
    
    # Alignment configs
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Fill patterns
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    # Borders
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Set sheet title row
    ws.merge_cells('A1:I1')
    ws['A1'] = "Apex Scaife Processing Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = left_align
    ws.row_dimensions[1].height = 40
    
    # Set Export date metadata info
    ws['A2'] = f"Generated On: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=9, italic=True)
    ws.row_dimensions[2].height = 20
    
    headers = [
        "Serial Number", "Client Name", "Diameter (in)", "Lapping Tier", 
        "Coating Tier", "Assigned Office", "Received Date", "Return Date", "Cost (INR)"
    ]
    
    # Write Headers
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 26

    # Write Data rows
    row_idx = 5
    total_cost = Decimal('0.00')
    
    for entry in queryset:
        ws.cell(row=row_idx, column=1, value=entry.serial_number).alignment = center_align
        ws.cell(row=row_idx, column=2, value=entry.client_name).alignment = left_align
        ws.cell(row=row_idx, column=3, value=float(entry.diameter)).alignment = right_align
        ws.cell(row=row_idx, column=4, value=entry.get_lapping_type_display()).alignment = left_align
        ws.cell(row=row_idx, column=5, value=entry.get_coating_type_display()).alignment = left_align
        ws.cell(row=row_idx, column=6, value=entry.assigned_office.name).alignment = left_align
        ws.cell(row=row_idx, column=7, value=entry.received_date.strftime("%d %b %Y")).alignment = center_align
        ws.cell(row=row_idx, column=8, value=entry.return_date.strftime("%d %b %Y") if entry.return_date else "-").alignment = center_align
        
        cost_cell = ws.cell(row=row_idx, column=9, value=float(entry.cost))
        cost_cell.alignment = right_align
        cost_cell.number_format = '₹#,##0.00'
        
        total_cost += entry.cost
        
        # Style row cells
        for col_idx in range(1, 10):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = regular_font
            c.border = thin_border
            if row_idx % 2 == 0:
                c.fill = zebra_fill
                
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    # Write Totals Row
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    total_label = ws.cell(row=row_idx, column=1, value="Grand Total Cost:")
    total_label.font = bold_font
    total_label.alignment = right_align
    total_label.border = thin_border
    
    total_val = ws.cell(row=row_idx, column=9, value=float(total_cost))
    total_val.font = bold_font
    total_val.alignment = right_align
    total_val.number_format = '₹#,##0.00'
    total_val.border = thin_border
    ws.row_dimensions[row_idx].height = 24

    # Auto fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        # Avoid checking merged cells to prevent column expansion bloating
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row > 2 and len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Return Excel download response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=scaife_report_{datetime.date.today().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    log_action(request.user, "REPORT", "Excel", "Scaife list", "Excel registry report exported.")
    return response


@login_required
def export_pdf(request):
    queryset = get_filtered_scaifes(request)
    branding = BrandingSettings.get_solo()
    
    # Create PDF response container
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=scaife_report_{datetime.date.today().strftime("%Y%m%d")}.pdf'
    
    # Document Setup (A4 landscape fits tables beautifully)
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Typography Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=6
    )
    
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=20
    )
    
    table_text = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#1E293B')
    )
    
    table_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1 # Center align
    )

    story = []
    
    # Header Letterhead
    story.append(Paragraph(branding.company_name, title_style))
    story.append(Paragraph(f"Scaife Processing Registry Audit Report — Generated on {datetime.datetime.now().strftime('%d %b %Y at %H:%M')}", meta_style))
    
    # Table Header data
    table_data = [[
        Paragraph("Serial No", table_header),
        Paragraph("Client Name", table_header),
        Paragraph("Dia (in)", table_header),
        Paragraph("Lapping Tier", table_header),
        Paragraph("Coating Tier", table_header),
        Paragraph("Office", table_header),
        Paragraph("Rec. Date", table_header),
        Paragraph("Cost", table_header),
    ]]
    
    # Add queryset values
    total_cost = Decimal('0.00')
    for entry in queryset:
        table_data.append([
            Paragraph(entry.serial_number, table_text),
            Paragraph(entry.client_name, table_text),
            Paragraph(f'{entry.diameter}"', table_text),
            Paragraph(entry.get_lapping_type_display(), table_text),
            Paragraph(entry.get_coating_type_display(), table_text),
            Paragraph(entry.assigned_office.name, table_text),
            Paragraph(entry.received_date.strftime("%d %b %Y"), table_text),
            Paragraph(f"INR {entry.cost}", table_text),
        ])
        total_cost += entry.cost

    # Add total row
    table_data.append([
        Paragraph("<b>Grand Total:</b>", table_text),
        "", "", "", "", "", "",
        Paragraph(f"<b>INR {total_cost}</b>", table_text)
    ])
    
    # Design Table dimensions (A4 printable width is ~535pt)
    # 8 columns distribution:
    col_widths = [65, 95, 40, 75, 75, 75, 55, 55]
    
    t = Table(table_data, colWidths=col_widths)
    
    # Style Table representation
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        
        # Grid lines
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#E2E8F0')),
        
        # Totals Row styles
        ('SPAN', (0,-1), (6,-1)),
        ('ALIGN', (0,-1), (0,-1), 'RIGHT'),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.HexColor('#1E293B')),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOTTOMPADDING', (0,-1), (-1,-1), 8),
        ('TOPPADDING', (0,-1), (-1,-1), 8),
    ]))
    
    # Zebra stripes styling loop
    for i in range(1, len(queryset) + 1):
        if i % 2 == 0:
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,i), (-1,i), colors.HexColor('#F8FAFC'))
            ]))
            
    story.append(t)
    doc.build(story)
    
    log_action(request.user, "REPORT", "PDF", "Scaife list", "PDF registry report exported.")
    return response


# ==============================================================
# WORKER PAYROLL & FINANCIAL LEDGER VIEWS
# ==============================================================

@login_required
@manager_or_admin
def worker_list(request):
    user_profile = request.user.profile
    if user_profile.role == 'admin':
        workers = WorkerProfile.objects.all().order_by('name')
    else:
        workers = WorkerProfile.objects.filter(office=user_profile.office).order_by('name')

    query = request.GET.get('q', '')
    filter_office = request.GET.get('office', '')

    if query:
        workers = workers.filter(name__icontains=query) | workers.filter(phone__icontains=query)
    if filter_office:
        workers = workers.filter(office_id=filter_office)

    offices = Office.objects.all()
    from decimal import Decimal
    context = {
        'workers': workers,
        'offices': offices,
        'query': query,
        'filter_office': filter_office,
        'total_workers_count': workers.count(),
        'grand_total_base_salary': sum((w.base_salary for w in workers), Decimal('0')),
        'grand_total_advances': sum((w.total_advance() for w in workers), Decimal('0')),
        'grand_total_payments': sum((w.total_payments() for w in workers), Decimal('0')),
        'grand_total_pending': sum((w.pending_balance() for w in workers), Decimal('0')),
    }
    return render(request, 'dashboard/worker_list.html', context)


@login_required
@admin_only
def worker_create(request):
    if request.method == 'POST':
        form = WorkerProfileForm(request.POST)
        if form.is_valid():
            worker = form.save()
            log_action(request.user, "CREATE", "WorkerProfile", worker.name,
                       f"Worker '{worker.name}' created with base salary ₹{worker.base_salary}.")
            messages.success(request, f"Worker '{worker.name}' created successfully.")
            return redirect('worker_detail', pk=worker.pk)
    else:
        form = WorkerProfileForm()
    return render(request, 'dashboard/worker_form.html', {'form': form, 'offices': Office.objects.all()})


@login_required
@manager_or_admin
def worker_detail(request, pk):
    worker = get_object_or_404(WorkerProfile, pk=pk)
    user_profile = request.user.profile
    if user_profile.role != 'admin' and worker.office != user_profile.office:
        messages.error(request, "You do not have permission to view workers from other offices.")
        return redirect('worker_list')

    today = datetime.date.today()
    filter_month = request.GET.get('month', str(today.month))
    filter_year  = request.GET.get('year',  str(today.year))

    qs = worker.financial_entries.all().order_by('-date', '-created_at')
    if filter_year  != 'all': qs = qs.filter(date__year=int(filter_year))
    if filter_month != 'all': qs = qs.filter(date__month=int(filter_month))

    from decimal import Decimal
    def stype(e): return sum((x.amount for x in qs.filter(entry_type=e)), Decimal('0.00'))

    sal = Decimal(str(worker.base_salary))
    adv = stype('advance')
    cash = stype('cash_taken')
    leave_days = sum(e.leave_days for e in qs.filter(entry_type='leave'))
    bonus = stype('bonus')
    ded = stype('deduction')
    exp = stype('expense')
    pay = stype('payment')

    # Total Taken/Settled by worker (Advance + Cash Taken + Direct Payment - Deduction)
    total_taken = (adv + cash + pay) - ded
    if total_taken < Decimal('0.00'):
        total_taken = Decimal('0.00')

    # Remaining pending salary:
    # Pending = (Salary + Bonus + Expense) - Deductions - (Advance + Cash Taken + Payment)
    pending = (sal + bonus + exp) - ded - (adv + cash + pay)

    metrics = {
        'salary':          sal,
        'advance':         adv,
        'cash_taken':      cash,
        'leave':           leave_days,
        'bonus':           bonus,
        'deduction':       ded,
        'expense':         exp,
        'payment':         total_taken,        # Total money already taken/settled (e.g. ₹7,500)
        'gross_taken':     adv + cash + pay,   # Total payouts before deduction
        'direct_payment':  pay,
        'pending':         pending,            # Remaining balance (e.g. ₹3,500 in red)
        'all_time_pending': worker.pending_balance(),
    }

    available_years = sorted({today.year} | {d.year for d in worker.financial_entries.dates('date', 'year')}, reverse=True)
    months_list = [
        ('1','January'),('2','February'),('3','March'),('4','April'),
        ('5','May'),('6','June'),('7','July'),('8','August'),
        ('9','September'),('10','October'),('11','November'),('12','December'),
    ]

    context = {
        'worker': worker, 'entries': qs, 'metrics': metrics,
        'entry_form': WorkerFinancialEntryForm(),
        'filter_month': filter_month, 'filter_year': filter_year,
        'available_years': available_years, 'months_list': months_list,
    }
    return render(request, 'dashboard/worker_detail.html', context)


@login_required
@admin_only
def worker_salary_update(request, pk):
    worker = get_object_or_404(WorkerProfile, pk=pk)
    if request.method == 'POST':
        new_salary = request.POST.get('base_salary')
        if new_salary is not None:
            try:
                from decimal import Decimal
                val = Decimal(str(new_salary).strip())
                if val >= Decimal('0.00'):
                    old_salary = worker.base_salary
                    worker.base_salary = val
                    worker.save()
                    log_action(request.user, "UPDATE", "WorkerProfile", worker.name,
                               f"Updated base salary for '{worker.name}' from ₹{old_salary} to ₹{val}.")
                    messages.success(request, f"Worker '{worker.name}' base salary updated to ₹{val:,.2f} successfully.")
                else:
                    messages.error(request, "Salary amount must be a positive number.")
            except Exception as e:
                messages.error(request, f"Invalid salary value: {e}")
    return redirect('worker_detail', pk=worker.pk)


@login_required
@admin_only
def worker_edit(request, pk):
    worker = get_object_or_404(WorkerProfile, pk=pk)
    if request.method == 'POST':
        form = WorkerProfileForm(request.POST, instance=worker)
        if form.is_valid():
            form.save()
            log_action(request.user, "UPDATE", "WorkerProfile", worker.name,
                       f"Worker '{worker.name}' profile updated.")
            messages.success(request, f"Worker '{worker.name}' updated successfully.")
            return redirect('worker_detail', pk=worker.pk)
    else:
        form = WorkerProfileForm(instance=worker)
    return render(request, 'dashboard/worker_form.html', {'form': form, 'worker': worker, 'offices': Office.objects.all()})


@login_required
@manager_or_admin
def worker_entry_add(request, pk):
    worker = get_object_or_404(WorkerProfile, pk=pk)
    if request.method == 'POST':
        form = WorkerFinancialEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.worker = worker
            entry.created_by = request.user
            entry.save()
            log_action(request.user, "CREATE", "WorkerFinancialEntry",
                       f"{worker.name} - {entry.get_entry_type_display()}",
                       f"Added {entry.get_entry_type_display()} entry ₹{entry.amount}.")
            messages.success(request, f"{entry.get_entry_type_display()} recorded for {worker.name}.")
        else:
            messages.error(request, "Failed to record entry. Please check input values.")
    return redirect('worker_detail', pk=worker.pk)


@login_required
@admin_only
def worker_entry_delete(request, pk):
    entry = get_object_or_404(WorkerFinancialEntry, pk=pk)
    worker_pk = entry.worker.pk
    log_action(request.user, "DELETE", "WorkerFinancialEntry",
               f"{entry.worker.name} - {entry.get_entry_type_display()}",
               f"Deleted transaction entry for {entry.worker.name}.")
    entry.delete()
    messages.success(request, "Transaction entry deleted.")
    return redirect('worker_detail', pk=worker_pk)


# ==========================================================================
# MASTER SECURITY PIN VERIFICATION & PWA MOBILE APP VIEWS
# ==========================================================================
from django.http import JsonResponse

@login_required
def verify_security_pin(request):
    """Verify admin master security PIN to reveal office manager passwords."""
    if request.method == 'POST':
        # Support both JSON body (fetch API) and regular form POST
        entered_pin = ''
        content_type = request.content_type or ''
        if 'application/json' in content_type:
            try:
                import json as _json
                body = _json.loads(request.body)
                entered_pin = str(body.get('pin', '')).strip()
            except Exception:
                entered_pin = ''
        else:
            entered_pin = (request.POST.get('pin') or '').strip()

        branding = BrandingSettings.get_solo()
        correct_pin = (branding.admin_security_pin or '1234').strip()

        if entered_pin and entered_pin == correct_pin:
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid Security PIN. Access Denied.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def pwa_manifest(request):
    """Return Web App Manifest JSON for PWA & Play/App Store compatibility."""
    branding = BrandingSettings.get_solo()
    
    icon_src = "/static/images/icon-512.png"
    icon_type = "image/png"
    if branding.logo:
        icon_src = branding.logo.url
        icon_type = "image/png" if branding.logo.name.lower().endswith('.png') else "image/jpeg"
    elif branding.favicon:
        icon_src = branding.favicon.url
        icon_type = "image/png" if branding.favicon.name.lower().endswith('.png') else "image/jpeg"

    manifest = {
        "name": branding.company_name,
        "short_name": "Scaife Portal",
        "description": branding.tagline,
        "start_url": "/",
        "display": "standalone",
        "background_color": "#0f172a",
        "theme_color": "#0c4a6e",
        "orientation": "any",
        "icons": [
            {
                "src": icon_src,
                "sizes": "512x512",
                "type": icon_type
            }
        ]
    }
    return JsonResponse(manifest)


def pwa_serviceworker(request):
    """Return Service Worker JS script for offline support & app installation."""
    sw_script = """
const CACHE_NAME = 'scaife-portal-v1';
const urlsToCache = ['/', '/dashboard/', '/static/css/styles.css', '/static/js/main.js'];

self.addEventListener('install', event => {
    event.waitUntil(
        caches.open(CACHE_NAME).then(cache => cache.addAll(urlsToCache))
    );
});

self.addEventListener('fetch', event => {
    event.respondWith(
        fetch(event.request).catch(() => caches.match(event.request))
    );
});
"""
    return HttpResponse(sw_script, content_type='application/javascript')

